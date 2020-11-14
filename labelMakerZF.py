import os, csv
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment


#Settings for main folder
superFolder = "C:\\Users\\user\\Downloads"
os.chdir(superFolder)

#User input to find range of rows to convert into labels
while True:
    try:
        startingRow = int(input("What row contains the first label: "))
        endingRow = int(input("What row contains the last label: "))
        print("Program starting up!")
        break
    except ValueError:
        print("Something you entered wasn't a digit!")

#Code to create/name an excel file
def makeExcel():
    global label
    global outputSheet
    label = Workbook()
    outputSheet = label.active

zfFont = Font(size = 16, bold = True)
lineFont = Font(size = 14, bold = True)
zygFont = Font(size = 10)
xFont = Font(size = 10, bold = True)
normalFont = Font(size = 8)
smallFont = Font(size = 5)
centerAlign = Alignment(horizontal = "center", vertical = "top")

#Code collect data and fill out data with formatting
def collectAndFill(fileName):
    inputExcel = load_workbook(fileName)
    inputSheet = inputExcel.active
    for row in range(startingRow, endingRow + 1):
        makeExcel()
        
        zfNumber = inputSheet.cell(row, 2).value
        line = inputSheet.cell(row, 3).value
        zygosity = inputSheet.cell(row, 4).value
        xNumber = inputSheet.cell(row, 7).value
        basicDOB = inputSheet.cell(row, 5).value
        dob = str(basicDOB).split(" ")[0]
        acup = inputSheet.cell(row, 9).value
        pi = inputSheet.cell(row, 10).value

        outputSheet.cell(1,1).value = zfNumber
        outputSheet.cell(2,1).value = line
        outputSheet.cell(3,1).value = zygosity
        outputSheet.cell(4,1).value = xNumber
        outputSheet.cell(5,1).value = dob
        outputSheet.cell(6,1).value = acup
        outputSheet.cell(7,1).value = pi
        outputSheet.cell(8,1).value = "Put Spawn Sticker Here"

        outputSheet.column_dimensions["A"].width = 13
        outputSheet.row_dimensions[8].height = 45
        outputSheet.cell(1,1).font = zfFont
        outputSheet.cell(2,1).font = lineFont
        outputSheet.cell(3,1).font = zygFont
        outputSheet.cell(4,1).font = xFont
        outputSheet.cell(5,1).font = normalFont
        outputSheet.cell(6,1).font = normalFont
        outputSheet.cell(7,1).font = normalFont
        outputSheet.cell(8,1).font = smallFont

        for i in range(1,9):
            outputSheet.cell(i,1).alignment = centerAlign

        title = "ZF " + str(int(float(zfNumber))) + " Line " + str(line) + " Label.xlsx"
        label.save(title)

collectAndFill("SARL STOCK BOOK.xlsx")
