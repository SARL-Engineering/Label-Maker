import os, csv
from openpyxl import Workbook
from openpyxl import load_workbook


#Settings for main folder
superFolder = 
os.chdir(superFolder)

#User input to find range of rows to convert into labels
while True:
    try:
        startingRow = int(input("What row contains the first label: "))
        endingRow = int(input("What row contains the last label: "))
        print("Program starting up!")
        break
    except ValueError:
        print("Something you entered wasn't a digit!")

#Code to create/name an excel file
def makeExcel():
    global label
    global outputSheet
    label = Workbook("LabelSheet")
    label.create_sheet("Label")
    outputSheet = label.active

#Code collect data and fill out data with formatting
def collectAndFill(fileName):
    inputExcel = load_workbook(fileName)
    inputSheet = inputExcel.active
    for row in range(startingRow, endingRow + 1):
        xxx = inputSheet.cell(row, 1).value

        outputSheet.cell(1,1).value = xxx

        title = "Row " + str(row) + " Label.xlsx"
        label.save(title)
